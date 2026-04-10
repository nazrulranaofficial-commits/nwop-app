import streamlit as st
import pandas as pd
import re
import time
import hashlib
import json
import os
import pytz
import uuid
import requests
import warnings
from io import BytesIO
from datetime import datetime, timedelta, date, time as dt_time
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from PIL import Image

# Suppress annoying Openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# --- SUPABASE IMPORT ---
try:
    from supabase import create_client, Client
    SUPABASE_AVAILABLE = True
except ImportError:
    SUPABASE_AVAILABLE = False

# --- SELENIUM IMPORTS ---
try:
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service as ChromeService
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.common.keys import Keys
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_AVAILABLE = True
except ImportError:
    SELENIUM_AVAILABLE = False

# --- AI IMPORTS ---
try:
    from google import genai
    GEMINI_AVAILABLE = True
except ImportError:
    GEMINI_AVAILABLE = False

try:
    from groq import Groq
    GROQ_AVAILABLE = True
except ImportError:
    GROQ_AVAILABLE = False

# --- SETUP ---
BD_TZ = pytz.timezone('Asia/Dhaka')
st.set_page_config(page_title="NWOP Pro", page_icon="🚀", layout="wide", initial_sidebar_state="expanded")

# --- GLOBAL HELPER FUNCTIONS ---
def get_image_bytes(filepath):
    if os.path.exists(filepath):
        try:
            with open(filepath, "rb") as img_file: return img_file.read()
        except: return None
    return None

def bn_to_en_digits(text): 
    return str(text).translate(str.maketrans('০১২৩৪৫৬৭৮৯', '0123456789'))

def format_phone_number(raw_phone):
    if not raw_phone: return "N/A"
    rp_str = str(raw_phone).strip().lower()
    if rp_str in ["n/a", "nan", ""]: return "N/A"
    clean = re.sub(r'\D', '', bn_to_en_digits(rp_str))
    if len(clean) < 10: return "N/A"
    if clean.startswith('880') and len(clean) > 11: clean = clean[2:] 
    elif clean.startswith('88') and len(clean) > 10: clean = clean[2:]
    if len(clean) == 10 and clean.startswith('1'): clean = '0' + clean 
    if len(clean) > 11: clean = clean[-11:]
    return clean if re.match(r'^01[3-9]\d{8}$', clean) else "N/A"

PHONE_PATTERN = r'((?:\+88|88)?0\s*1\s*[3-9](?:[\s-]*\d){8})'

def check_message_status(text_en): 
    digits = re.sub(r'\D', '', text_en)
    return "valid" if len(digits) >= 10 else "ignored"

def clean_system_messages(text):
    lines = str(text).split('\n')
    clean_lines = []
    patterns = [' added ', ' joined using', ' left', ' changed the subject', ' removed ', 'security code changed', 'encrypted']
    for line in lines:
        if not any(p in line.lower() for p in patterns):
            clean_lines.append(line)
    return '\n'.join(clean_lines)

def is_whatsapp_system_message(text):
    text_lower = str(text).lower()
    patterns = [' added ', ' joined using', ' left', ' changed the subject', ' removed ', 'security code changed', 'encrypted']
    return any(p in text_lower for p in patterns)

def get_datetime_obj(date_string, time_string):
    try:
        if ':' in date_string and '/' in time_string: date_string, time_string = time_string, date_string
        d_clean = date_string.strip()
        if len(d_clean.split('/')[-1]) == 2:
            try: d = datetime.strptime(d_clean, "%d/%m/%y").date()
            except: d = datetime.strptime(d_clean, "%m/%d/%y").date()
        else:
            try: d = datetime.strptime(d_clean, "%d/%m/%Y").date()
            except: d = datetime.strptime(d_clean, "%m/%d/%Y").date()
        t_clean = time_string.replace('\u202f', ' ').replace('\u200e', '').replace('\u200f', '').strip()
        if 'AM' not in t_clean.upper() and 'PM' not in t_clean.upper():
            if t_clean.count(':') == 2: t = datetime.strptime(t_clean, "%H:%M:%S").time()
            else: t = datetime.strptime(t_clean, "%H:%M").time()
        else:
            if t_clean.count(':') == 2: t = datetime.strptime(t_clean, "%I:%M:%S %p").time()
            else: t = datetime.strptime(t_clean, "%I:%M %p").time()
        return datetime.combine(d, t)
    except: return None

def parse_copy_paste_time(pasted_str):
    if not pasted_str: return None
    clean = pasted_str.replace('[', '').replace(']', '').replace('\u200e', '').replace('\u200f', '').strip()
    if ',' in clean:
        parts = clean.split(',', 1)
        return get_datetime_obj(parts[0].strip(), parts[1].strip())
    return None

# --- UI STYLE ---
st.markdown("""
    <style>
    html { scroll-behavior: smooth; }
    .stApp { background-color: #090B10 !important; font-family: 'Inter', sans-serif; color: #E0E0E0; }
    .block-container { padding-top: 2rem !important; padding-bottom: 3rem !important; max-width: 1100px !important; }
    [data-testid="stToolbar"] a { display: none !important; } footer { display: none !important; }

    [data-testid="stExpander"] { background: rgba(255, 255, 255, 0.04) !important; backdrop-filter: blur(18px) !important; border-radius: 18px !important; border: 1px solid rgba(255, 255, 255, 0.08) !important; box-shadow: 0 10px 30px 0 rgba(0, 0, 0, 0.3) !important; margin-bottom: 18px !important; transition: all 0.3s ease-in-out; }
    [data-testid="stExpander"]:hover { box-shadow: 0 10px 30px 0 rgba(16, 185, 129, 0.15) !important; border: 1px solid rgba(16, 185, 129, 0.25) !important; }
    [data-testid="stExpander"] > details > summary { padding: 18px !important; font-weight: 800 !important; font-size: 1.1rem !important; color: #FFFFFF !important; }
    
    [data-testid="stMetric"] { background: rgba(255, 255, 255, 0.04) !important; backdrop-filter: blur(16px) !important; border-radius: 20px !important; padding: 20px 10px !important; text-align: center !important; border: 1px solid rgba(255, 255, 255, 0.08) !important; box-shadow: 0 4px 15px rgba(0, 0, 0, 0.3) !important; }
    [data-testid="stMetricValue"] { font-size: 2.2rem !important; font-weight: 900 !important; background: linear-gradient(135deg, #10B981, #3B82F6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; }
    
    .stButton > button { border-radius: 25px !important; border: 1px solid rgba(255, 255, 255, 0.15) !important; padding: 10px 20px !important; font-weight: 800 !important; letter-spacing: 0.5px !important; width: 100% !important; background: rgba(255, 255, 255, 0.05) !important; color: #FFFFFF !important; backdrop-filter: blur(10px) !important; transition: all 0.3s ease !important; }
    .stButton > button:hover { transform: translateY(-3px) !important; box-shadow: 0 8px 20px rgba(16, 185, 129, 0.25) !important; border: 1px solid #10B981 !important; }
    .stButton > button[kind="primary"] { background: linear-gradient(135deg, #10B981, #059669) !important; color: white !important; border: none !important; box-shadow: 0 5px 15px rgba(16, 185, 129, 0.4) !important; }

    .doubt-card { padding: 15px 20px; border-left: 5px solid #EF4444; background: rgba(239, 68, 68, 0.1); border-radius: 12px; margin-bottom: 12px; display: flex; justify-content: space-between; align-items: center; box-shadow: 0 4px 15px rgba(0,0,0,0.15); border: 1px solid rgba(239, 68, 68, 0.3); color: white; }
    .doubt-card .fix-btn { background: linear-gradient(135deg, #EF4444, #DC2626); color: white !important; padding: 8px 18px; border-radius: 20px; text-decoration: none; font-weight: 800; font-size: 0.9rem; transition: 0.3s; box-shadow: 0 4px 10px rgba(239, 68, 68, 0.3); }

    .wa-bubble { background: linear-gradient(135deg, rgba(7, 94, 84, 0.95), rgba(18, 140, 126, 0.85)); backdrop-filter: blur(15px); border: 1px solid rgba(255, 255, 255, 0.2); color: white; padding: 18px 22px; border-radius: 20px; border-top-left-radius: 0px; max-width: 95%; box-shadow: 0 8px 25px rgba(0,0,0,0.4); margin-bottom: 25px; line-height: 1.6; }

    .stTabs [data-baseweb="tab-list"] { background-color: transparent !important; gap: 10px !important; padding-bottom: 15px !important; }
    .stTabs [data-baseweb="tab"] { background: rgba(255, 255, 255, 0.05) !important; border-radius: 30px !important; padding: 10px 25px !important; border: 1px solid rgba(255,255,255,0.1) !important; font-weight: 700 !important; font-size: 1rem !important; color: #E0E0E0 !important; }
    .stTabs [aria-selected="true"] { background: linear-gradient(135deg, #10B981, #059669) !important; color: #FFFFFF !important; border: none !important; box-shadow: 0 5px 15px rgba(16, 185, 129, 0.4) !important; }

    .stTextInput input, .stNumberInput input, .stSelectbox div[data-baseweb="select"] { border-radius: 12px !important; border: 1px solid rgba(255, 255, 255, 0.15) !important; padding: 12px !important; background: rgba(255, 255, 255, 0.05) !important; color: white !important; font-weight: 500 !important; }
    .stTextInput input:focus, .stNumberInput input:focus { border: 1px solid #10B981 !important; box-shadow: 0 0 8px rgba(16, 185, 129, 0.4) !important; background: rgba(0,0,0,0.4) !important; }

    .dev-badge { margin-top: 30px; padding-top: 20px; border-top: 1px solid rgba(255, 255, 255, 0.1); font-size: 0.85rem; color: #A0A0A0; text-align: center;}
    .stChatMessage { background: rgba(255,255,255,0.03) !important; border-radius: 15px !important; padding: 15px !important; margin-bottom: 10px !important; border: 1px solid rgba(255,255,255,0.05) !important; }

    @media (max-width: 768px) { .stTabs [data-baseweb="tab-list"] { overflow-x: auto; white-space: nowrap; flex-wrap: nowrap; } .doubt-card { flex-direction: column; align-items: flex-start; gap: 15px; } }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 🔐 LOCAL AUTH VAULT (AUTO-SAVE CREDENTIALS)
# ==========================================
AUTH_VAULT_FILE = "nwop_auth.json"

def load_local_auth():
    if os.path.exists(AUTH_VAULT_FILE):
        try:
            with open(AUTH_VAULT_FILE, "r") as f: return json.load(f)
        except: return {}
    return {}

def save_local_auth(url, key, email, password):
    try:
        with open(AUTH_VAULT_FILE, "w") as f:
            json.dump({"sb_url": url, "sb_key": key, "sb_email": email, "sb_password": password}, f)
    except: pass

auth_data = load_local_auth()
if 'sb_url' not in st.session_state: st.session_state.sb_url = auth_data.get('sb_url', '')
if 'sb_key' not in st.session_state: st.session_state.sb_key = auth_data.get('sb_key', '')
if 'sb_email' not in st.session_state: st.session_state.sb_email = auth_data.get('sb_email', '')
if 'sb_password' not in st.session_state: st.session_state.sb_password = auth_data.get('sb_password', '')

# ==========================================
# 🔐 AUTO-LOGIN VIA STREAMLIT SECRETS 🔐
# ==========================================
if 'logged_in' not in st.session_state: 
    st.session_state.logged_in = False
    if SUPABASE_AVAILABLE:
        try:
            if hasattr(st, "secrets") and "SUPABASE_URL" in st.secrets and "SUPABASE_KEY" in st.secrets:
                supabase: Client = create_client(st.secrets["SUPABASE_URL"], st.secrets["SUPABASE_KEY"])
                res = supabase.auth.sign_in_with_password({
                    "email": st.secrets.get("SUPABASE_EMAIL", ""), 
                    "password": st.secrets.get("SUPABASE_PASSWORD", "")
                })
                if res.user:
                    st.session_state.supabase = supabase
                    st.session_state.user = res.user
                    st.session_state.logged_in = True
        except Exception: pass

if not SUPABASE_AVAILABLE:
    st.error("⚠️ Supabase is not installed! Run: `pip install supabase` in your terminal.")
    st.stop()

# 🌟 CLEAN NATIVE LOGIN SCREEN 🌟
if not st.session_state.logged_in:
    st.markdown("<br><br><br>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 1.5, 1])
    with col:
        with st.container(border=True):
            st.markdown("<div style='text-align: center;'>", unsafe_allow_html=True)
            img_bytes = get_image_bytes("logo.png")
            if img_bytes: st.image(img_bytes, width=120)
            else: st.markdown("<h1 style='font-size: 60px; margin: 0;'>🚀</h1>", unsafe_allow_html=True)
            st.markdown("<h2 style='font-weight:900; background: linear-gradient(135deg, #10B981, #3B82F6); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom:0;'>NWOP CLOUD</h2>", unsafe_allow_html=True)
            st.markdown("<p style='color: #A0A0A0; font-size: 14px;'>Enterprise Authentication</p>", unsafe_allow_html=True)
            st.markdown("</div><br>", unsafe_allow_html=True)
            
            in_sb_url = st.text_input("Supabase Project URL", value=st.session_state.sb_url, placeholder="https://xyz.supabase.co")
            in_sb_key = st.text_input("Supabase Anon Key", type="password", value=st.session_state.sb_key, placeholder="ey...")
            in_email = st.text_input("Email", value=st.session_state.sb_email, placeholder="admin@domain.com")
            in_password = st.text_input("Password", type="password", value=st.session_state.sb_password, placeholder="••••••••")
            st.write("")
            
            if st.button("Unlock Dashboard", type="primary", use_container_width=True):
                if not in_sb_url or not in_sb_key or not in_email or not in_password:
                    st.error("❌ All fields are required!")
                else:
                    try:
                        supabase: Client = create_client(in_sb_url, in_sb_key)
                        res = supabase.auth.sign_in_with_password({"email": in_email, "password": in_password})
                        if res.user:
                            save_local_auth(in_sb_url, in_sb_key, in_email, in_password)
                            st.session_state.supabase = supabase
                            st.session_state.user = res.user
                            st.session_state.sb_url = in_sb_url
                            st.session_state.sb_key = in_sb_key
                            st.session_state.sb_email = in_email
                            st.session_state.sb_password = in_password
                            st.session_state.logged_in = True
                            st.rerun()
                    except Exception as e:
                        st.error(f"❌ Auth Failed: {e}")
                
            st.markdown("<div class='dev-badge'>Developed by <b>Nazrul Rana</b><br>v53.0 The 100% Masterpiece</div>", unsafe_allow_html=True)
    st.stop()

# ==========================================
# ☁️ SUPABASE DATA SYNC LOGIC ☁️
# ==========================================
DEFAULT_PRODUCTS = ['Silver Crest Electric Blender', 'Electronic Grinder', 'Electric Blender', 'Vita Gold', 'Rice Cooker', 'Sound Box', 'Nima Blender', 'E-9 Pro', 'Self Stick', 'Shoe Rack', 'Light', 'Sky', 'Black', 'Rack', 'Green', 'Pink', 'Navy', 'Cream', 'Olive', 'White', 'Bottle', 'Check Manually']

def load_supabase_profile():
    try:
        res = st.session_state.supabase.table('nwop_profiles').select('settings').eq('user_id', st.session_state.user.id).execute()
        if res.data: return res.data[0]['settings']
        else:
            def_settings = {
                "history": [], "last_checkpoint": "No record yet", "groq_api_key": "", "gemini_api_key": "",
                "pathao_client_id": "", "pathao_client_secret": "", "pathao_store_id": "", "pathao_email": "", 
                "pathao_password": "", "learned_products": []
            }
            st.session_state.supabase.table('nwop_profiles').insert({'user_id': st.session_state.user.id, 'settings': def_settings}).execute()
            return def_settings
    except Exception as e:
        return {}

def save_supabase_profile(settings_dict):
    try:
        st.session_state.supabase.table('nwop_profiles').update({'settings': settings_dict}).eq('user_id', st.session_state.user.id).execute()
    except Exception as e: pass

def push_order_to_supabase(order_dict):
    try:
        payload = {"user_id": st.session_state.user.id, "order_data": order_dict}
        st.session_state.supabase.table('nwop_orders').insert(payload).execute()
    except: pass

def auto_delete_90_days_data():
    try:
        ninety_days_ago = (datetime.now(BD_TZ) - timedelta(days=90)).isoformat()
        st.session_state.supabase.table('nwop_orders').delete().lt('created_at', ninety_days_ago).execute()
    except: pass

if 'profile_loaded' not in st.session_state:
    auto_delete_90_days_data()
    prof = load_supabase_profile()
    st.session_state.task_history = prof.get("history", [])
    st.session_state.last_checkpoint = prof.get("last_checkpoint", "No record yet")
    st.session_state.groq_api_key = prof.get("groq_api_key", "")
    st.session_state.gemini_api_key = prof.get("gemini_api_key", "")
    st.session_state.pathao_client_id = prof.get("pathao_client_id", "")
    st.session_state.pathao_client_secret = prof.get("pathao_client_secret", "")
    st.session_state.pathao_store_id = prof.get("pathao_store_id", "")
    st.session_state.pathao_email = prof.get("pathao_email", "")
    st.session_state.pathao_password = prof.get("pathao_password", "")
    st.session_state.product_list = list(dict.fromkeys(DEFAULT_PRODUCTS + prof.get("learned_products", [])))
    st.session_state.profile_loaded = True

if 'analyze_engine' not in st.session_state: st.session_state.analyze_engine = "Groq: Llama 3.1 8B (Fast & Reliable)"
if 'chat_history' not in st.session_state: st.session_state.chat_history = [{"role": "assistant", "content": "হ্যালো! আমি NWOP এআই। আপনার বিজনেস ডাটা বা অন্য যেকোনো বিষয়ে আমি সাহায্য করতে প্রস্তুত।"}]
if 'all_orders' not in st.session_state: st.session_state.all_orders = []
if 'ignored_messages' not in st.session_state: st.session_state.ignored_messages = []
if 'total_scanned' not in st.session_state: st.session_state.total_scanned = 0
if 'sheet_date' not in st.session_state: st.session_state.sheet_date = datetime.now(BD_TZ).strftime("%d/%m/%y")
if 'total_extracted_today' not in st.session_state: st.session_state.total_extracted_today = 0
if 'is_sending_bulk' not in st.session_state: st.session_state.is_sending_bulk = False

def sync_profile_to_db():
    learned = [p for p in st.session_state.product_list if p not in DEFAULT_PRODUCTS]
    settings_dict = {
        "history": st.session_state.task_history, "last_checkpoint": st.session_state.last_checkpoint,
        "groq_api_key": st.session_state.groq_api_key, "gemini_api_key": st.session_state.gemini_api_key,
        "pathao_client_id": st.session_state.pathao_client_id, "pathao_client_secret": st.session_state.pathao_client_secret,
        "pathao_store_id": st.session_state.pathao_store_id, "pathao_email": st.session_state.pathao_email,
        "pathao_password": st.session_state.pathao_password, "learned_products": learned
    }
    save_supabase_profile(settings_dict)

def learn_new_product(product_name):
    if product_name and product_name not in st.session_state.product_list:
        st.session_state.product_list.append(product_name)
        sync_profile_to_db()

def log_task(task_desc):
    timestamp = datetime.now(BD_TZ).strftime("%d %b %Y, %I:%M %p")
    st.session_state.task_history.insert(0, f"✅ **{timestamp}**: {task_desc}")
    sync_profile_to_db()

# 🌟 BASE REGEX EXTRACTOR (ZERO DROP EDITION) 🌟
def extract_order_details(msg_dict):
    text = msg_dict["text"]
    raw_text = text  
    parts = re.split(r'^\[.*?\] .*?:\s', text, maxsplit=1)
    body = parts[1] if len(parts) > 1 else text
    
    if is_whatsapp_system_message(body):
        return {"status": "ignored", "Date": msg_dict["date_str"], "Time": msg_dict["time_str"], "Text": raw_text, "Reason": "Auto-generated System Message", "id": str(uuid.uuid4())}
    
    body_en = bn_to_en_digits(body)
    body_en = re.sub(r'(\d),(\d)', r'\1\2', body_en)

    status = check_message_status(body_en)
    if status == "ignored":
        return {"status": "ignored", "Date": msg_dict["date_str"], "Time": msg_dict["time_str"], "Text": raw_text, "Reason": "Valid Phone Number Missing", "id": str(uuid.uuid4())}

    body_en = body_en.replace('<This message was edited>', ' ')
    body_en = re.sub(r'অর্ডার\s*করতে\s*[-ঃ:]*\s*', ' ', body_en, flags=re.IGNORECASE)

    phone_match = re.search(PHONE_PATTERN, body_en)
    phone = "N/A"
    if phone_match:
        raw_phone = phone_match.group(1)
        body_en = body_en.replace(raw_phone, ' ') 
        phone = format_phone_number(raw_phone)

    price = 0
    eq_match = re.search(r'\d+\s*\+\s*\d+\s*=\s*(\d+)', body_en)
    if eq_match:
        price = int(eq_match.group(1))
        body_en = body_en.replace(eq_match.group(0), ' ') 
    else:
        price_match = re.search(r'(\d+)\s*(?:টাকা|taka|tk|/-)', body_en, re.IGNORECASE)
        if price_match:
            price = int(price_match.group(1))
            body_en = body_en.replace(price_match.group(0), ' ') 
        else:
            price_match = re.search(r'=\s*(\d+)', body_en)
            if price_match:
                price = int(price_match.group(1))
                body_en = body_en.replace(price_match.group(0), ' ') 

    qty_match = re.search(r'(\d+)\s*(?:pcs?|pis|পিস|piece|টি|টা|ta(?!\w))', body_en, re.IGNORECASE)
    quantity = int(qty_match.group(1)) if qty_match else 1
    if qty_match: body_en = body_en.replace(qty_match.group(0), ' ')

    product = "Electric Blender"
    if not re.search(r'grind|grainder|blender', body_en, re.IGNORECASE): pass 

    body_en = re.sub(r'electrc|electric|electronic|blenders?|grinders?|grainders?|food\s*grind|taka|tk|টাকা', ' ', body_en, flags=re.IGNORECASE)
    body_en = body_en.replace('/-', ' ')
    body_en = re.sub(r'image omitted|<media omitted>|media omitted', ' ', body_en, flags=re.IGNORECASE)

    clean_body = body_en.replace('\n', ',').replace('=', ',').replace('।।', ',').replace('।', ',').replace('|', ',')
    clean_body = re.sub(r'(?i)(?<![a-zA-Z])dist?\.', 'জেলা ', clean_body)
    
    major_labels = ['নাম', 'name', 'nam', 'ফুল ঠিকানা', 'ঠিকানা', 'ঠীকানা', 'thikana', 'address', 'add', 'এড্রেস']
    for kw in major_labels: clean_body = re.sub(rf'(?<![a-zA-Z0-9\u0980-\u09FF,])({kw})', r',\1', clean_body, flags=re.IGNORECASE)
        
    address_indicators = ['থানা', 'জেলা', 'গ্রাম', 'পোস্ট', 'বাজার', 'রোড', 'সদর', 'উপজেলা', 'মোড়', 'para', 'pur', 'gram', 'thana', 'bazar', 'road', 'zilla', 'district', 'upazila', 'বিভাগ', 'ওয়ার্ড', 'ঢাকা','চট্টগ্রাম','রাজশাহী','খুলনা','বরিশাল','সিলেট','রংপুর','ময়মনসিংহ','কুমিল্লা','নোয়াখালী','ফেনী','চাঁদপুর','ব্রাহ্মণবাড়িয়া','গাজীপুর','টাঙ্গাইল','নারায়ণগঞ্জ','নরসিংদী','ফরিদপুর','মাদারীপুর','শরীয়তপুর','গোপালগঞ্জ','কিশোরগঞ্জ','সুনামগঞ্জ','হবিগঞ্জ','মৌলভীবাজার','রাঙ্গামাটি','বান্দরবান','খাগড়াছড়ি','কক্সবাজার','লক্ষ্মীপুর','ভোলা','পটুয়াখালী','বরগুনা','ঝালকাঠি','পিরোজপুর','যশোর','সাতক্ষীরা','ঝিনাইদহ','মাগুরা','নড়াইল','বাগেরহাট','কুষ্টিয়া','কুষ্টিয়া','চুয়াডাঙ্গা','মেহেরপুর','পাবনা','সিরাজগঞ্জ','বগুড়া','জয়পুরহাট','নওগাঁ','নাটোর','চাঁপাইনবাবগঞ্জ','দিনাজপুর','ঠাকুরগাঁও','পঞ্চগড়','নীলফামারী','কুড়িগ্রাম','লালমনিরহাট','গাইবান্ধা','জামালপুর','শেরপুর','নেত্রকোণা']
    for kw in address_indicators: clean_body = re.sub(rf'(?<![a-zA-Z0-9\u0980-\u09FF,])({kw})', r',\1', clean_body, flags=re.IGNORECASE)

    clean_body = re.sub(r'(?:মোবাইল\s*নাম্বার|মোবাইল|ফোন\s*নাম্বার|ফোন|নাম্বার|mobile\s*number|mobile|phone\s*number|phone|number)[\sঃ:=-]*', '', clean_body, flags=re.IGNORECASE)
    
    raw_chunks = [c.strip() for c in clean_body.split(',') if c.strip()]
    name, address_lines = "N/A", []
    explicit_name_found = False

    for chunk in raw_chunks:
        if not re.search(r'[a-zA-Zঅ-য়0-9]', chunk): continue
        cleaned_chunk = re.sub(r'^[+0-9\s-]+$', '', chunk).strip()
        if not cleaned_chunk: continue

        if re.match(r'^(?:নাম|name|nam)\s*[:ঃ=-]+\s*|^(?:নাম|name|nam)\s+', cleaned_chunk, re.IGNORECASE):
            name_val = re.sub(r'^(?:নাম|name|nam)\s*[:ঃ=-]*\s*', '', cleaned_chunk, flags=re.IGNORECASE).strip()
            if name_val:
                if name != "N/A" and not explicit_name_found: address_lines.insert(0, name) 
                name = name_val
                explicit_name_found = True
            continue
            
        if re.match(r'^(?:ফুল\s*ঠিকানা|ঠিকানা|ঠীকানা|thikana|address|add|এড্রেস)\s*[:ঃ=-]+\s*|^(?:ফুল\s*ঠিকানা|ঠিকানা|ঠীকানা|thikana|address|add|এড্রেস)\s+', cleaned_chunk, re.IGNORECASE):
            addr_val = re.sub(r'^(?:ফুল\s*ঠিকানা|ঠিকানা|ঠীকানা|thikana|address|add|এড্রেস)\s*[:ঃ=-]*\s*', '', cleaned_chunk, flags=re.IGNORECASE).strip()
            if addr_val: address_lines.append(addr_val)
            continue
            
        cleaned_chunk = re.sub(r'^(?:নাম|name|nam|ফুল\s*ঠিকানা|ঠিকানা|ঠীকানা|thikana|address|add|এড্রেস)\s*[:ঃ=-]*\s*', '', cleaned_chunk, flags=re.IGNORECASE).strip()
        if not cleaned_chunk: continue

        if name == "N/A": name = cleaned_chunk
        else: address_lines.append(cleaned_chunk)

    if name == "N/A" and address_lines: name = address_lines.pop(0)

    addr_hints = ['বাড়ি', 'বাড়ি', 'বাড়ী', 'বাড়ী', 'তলা', 'রোড', 'road', 'house', 'হাউজ', 'ফ্ল্যাট', 'flat', 'গ্রাম', 'থানা', 'জেলা', 'উপজেলা', 'মার্কেট', 'বটতলা', 'বাজার', 'কলেজ', 'গেট', 'gate', 'মোড়', 'mor', 'স্ট্যান্ড', 'stand', 'পাড়া', 'পাড়া', 'para', 'pur', 'পুর', 'নগর', 'nagar', 'ভবন', 'bhaban', 'building', 'tower', 'টাওয়ার', 'এলাকা', 'এভেনিউ', 'avenue', 'ব্লক', 'block', 'সেকশন', 'section', 'লেন', 'lane']
    
    if name != "N/A" and not explicit_name_found and len(address_lines) > 0:
        name_is_addr = any(hint.lower() in name.lower() for hint in addr_hints)
        if name_is_addr:
            for i in range(len(address_lines)-1, -1, -1):
                candidate = address_lines[i]
                if len(candidate.split()) <= 3 and not any(hint.lower() in candidate.lower() for hint in addr_hints):
                    real_name = candidate; address_lines.pop(i); address_lines.insert(0, name); name = real_name; break

    address = ", ".join(address_lines) if address_lines else "N/A"
    address = re.sub(r',+', ',', address); address = re.sub(r'\s*,\s*', ', ', address); address = address.strip(' ,-:;') 

    return {
        "id": str(uuid.uuid4()), "status": "valid", "Date": msg_dict["date_str"], "Time": msg_dict["time_str"],
        "Name": name, "Phone Number": phone, "Address": address, "Product": product,
        "Quantity": quantity, "Price": price, "Approval": "Pending", "Note": "", "is_duplicate": False,
        "RawText": raw_text, "Method": "⚙️ Regex"
    }

# 🌟 AI ANALYZE (SINGLE ORDER RECOVERY - GOOGLE.GENAI) 🌟
def analyze_single_order(raw_text, engine, groq_key, gem_key):
    prompt = f"""
    Extract order details from this RAW WhatsApp message perfectly.
    CRITICAL RULES:
    1. KEEP ORIGINAL LANGUAGE: If the Name or Address is in Bangla, you MUST keep it exactly in Bangla. DO NOT translate to English.
    2. Name: Extract the customer's name exactly as written.
    3. Address: Must contain the FULL address exactly as written. DO NOT omit any part! Remove labels like "add:".
    4. Phone Number: ONLY FOR PHONE NUMBER convert Bangla digits (০-৯) to English digits (0-9). Extract the raw phone number.
    5. Price: Find the final total. Ignore commas. Do not do math.
    6. Quantity: Extract integer. Default 1.
    7. Product: "Silver Crest" + "Blender" -> "Silver Crest Electric Blender". Or exact text.
    8. RETURN JSON FORMAT EXACTLY: {{"Name": "...", "Phone Number": "...", "Address": "...", "Product": "...", "Quantity": 1, "Price": 0}}

    Raw Message:
    {raw_text}
    """
    try:
        if "Groq" in engine:
            client = Groq(api_key=groq_key)
            model_n = "llama-3.3-70b-versatile" if "70B" in engine else "llama-3.1-8b-instant"
            response = client.chat.completions.create(messages=[{"role": "user", "content": prompt}], model=model_n, temperature=0.1, response_format={"type": "json_object"})
            return json.loads(response.choices[0].message.content)
        elif "Gemini" in engine:
            client = genai.Client(api_key=gem_key)
            response = client.models.generate_content(model='gemini-1.5-flash', contents=prompt)
            text = re.sub(r'^```json\s*|\s*```$', '', response.text.strip(), flags=re.IGNORECASE|re.MULTILINE).strip()
            match = re.search(r'\{[\s\S]*\}', text)
            if match: return json.loads(match.group(0))
    except Exception as e:
        print(f"AI Error: {e}")
        return None

# 🌟 AI VISION ENGINE (GOOGLE.GENAI SDK) 🌟
def extract_from_image_vision(image_file, api_key):
    try:
        client = genai.Client(api_key=api_key)
        img = Image.open(image_file)
        prompt = """
        Act as a strict OCR and Data Extractor for an E-commerce business.
        Read the provided image (handwritten or printed in Bangla/English).
        Extract the following and output ONLY a raw JSON object:
        { "Name": "Clean Name", "Phone Number": "01...", "Address": "Full Address", "Product": "Product Name", "Quantity": integer, "Price": integer }
        Do not add Markdown. Do not calculate math. Keep original language (Bangla) except phone numbers.
        """
        response = client.models.generate_content(model='gemini-1.5-flash', contents=[prompt, img])
        text = re.sub(r'^```json\s*|\s*```$', '', response.text.strip(), flags=re.IGNORECASE|re.MULTILINE).strip()
        text = re.sub(r'^```\s*|\s*```$', '', text, flags=re.IGNORECASE|re.MULTILINE).strip()
        match = re.search(r'\{[\s\S]*\}', text)
        if match: return json.loads(match.group(0))
    except Exception as e: 
        st.error(f"Vision API Error: {e}")
        return None

# 🌟 PATHAO API (BULLETPROOF INCOMPLETE ADDRESS DETECTOR) 🌟
def send_to_pathao_api(order_data, client_id, client_secret, store_id, email, password):
    try:
        token_url = "https://api-hermes.pathao.com/aladdin/api/v1/issue-token"
        token_payload = {"client_id": client_id, "client_secret": client_secret, "grant_type": "password", "username": email, "password": password}
        token_res = requests.post(token_url, json=token_payload)
        
        if token_res.status_code != 200: return False, f"Auth Failed: Check Client ID/Secret or Email/Password."
        access_token = token_res.json().get("access_token")
        
        order_url = "https://api-hermes.pathao.com/aladdin/api/v1/orders"
        headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/json", "Accept": "application/json"}
        
        original_addr = str(order_data.get("Address", "")).lower()
        payload = {
            "store_id": int(store_id),
            "merchant_order_id": str(order_data.get("id", str(uuid.uuid4())))[:15],
            "recipient_name": str(order_data.get("Name", "Customer"))[:50],
            "recipient_phone": str(order_data.get("Phone Number", ""))[:15],
            "recipient_address": str(order_data.get("Address", "N/A"))[:200],
            "delivery_type": 48,
            "item_type": 2, 
            "special_instruction": str(order_data.get("Note", ""))[:200],
            "item_quantity": int(order_data.get("Quantity", 1)), 
            "item_weight": 1, 
            "amount_to_collect": int(order_data.get("Price", 0))
        }
        order_res = requests.post(order_url, headers=headers, json=payload)
        
        if order_res.status_code in [200, 201]:
            resp_data = order_res.json().get('data', {})
            cons_id = resp_data.get('consignment_id', 'Success')
            
            is_incomplete = False
            raw_resp = order_res.text
            str_resp_lower = raw_resp.lower()
            
            # Heuristic 1: Exact City 1 + Zone 1 JSON format check
            if re.search(r'"city_id"\s*:\s*1\b', str_resp_lower) and re.search(r'"zone_id"\s*:\s*1\b', str_resp_lower):
                is_incomplete = True
            elif re.search(r'"recipient_city"\s*:\s*1\b', str_resp_lower) and re.search(r'"recipient_zone"\s*:\s*1\b', str_resp_lower):
                is_incomplete = True
                
            # Heuristic 2: The "Banani" Injection Trick (Bulletproof for API default fallback)
            if "banani" in str_resp_lower and "banani" not in original_addr and "বনানী" not in original_addr:
                is_incomplete = True
                
            if not is_incomplete and cons_id != 'Success':
                try:
                    info_res = requests.get(f"https://api-hermes.pathao.com/aladdin/api/v1/orders/{cons_id}", headers=headers, timeout=8)
                    raw_info = info_res.text.lower()
                    if re.search(r'"city_id"\s*:\s*1\b', raw_info) and re.search(r'"zone_id"\s*:\s*1\b', raw_info):
                        is_incomplete = True
                    elif re.search(r'"recipient_city"\s*:\s*1\b', raw_info) and re.search(r'"recipient_zone"\s*:\s*1\b', raw_info):
                        is_incomplete = True
                    elif "banani" in raw_info and "banani" not in original_addr and "বনানী" not in original_addr:
                        is_incomplete = True
                except: pass
                
            if is_incomplete:
                return "INCOMPLETE", f"{cons_id}"
            
            return True, f"{cons_id}"
        else: return False, f"Pathao Error: {order_res.json().get('message', 'Invalid Entry')}"
    except Exception as e: return False, f"Connection Error: {str(e)}"

# --- EXCEL GENERATOR HELPER ---
def generate_excel_bytes(orders_list, sheet_date, product_list):
    export_data = []
    for order in orders_list:
        clean_order = {k: v for k, v in order.items() if k not in ['is_duplicate', 'id', 'RawText', 'Expander_Title', 'Method', 'is_sent', 'temp_id', '❌ Drop']}
        export_data.append(clean_order)
        push_order_to_supabase(clean_order)
        
    export_df = pd.DataFrame(export_data)
    export_df['Quantity'] = pd.to_numeric(export_df.get('Quantity', 1), errors='coerce').fillna(1).astype(int)
    export_df['is_multi'] = export_df['Quantity'] > 1
    export_df = export_df.sort_values(by=['is_multi']).drop(columns=['is_multi'])
    export_df['SNO'] = range(1, 1 + len(export_df))
    
    csv_df = export_df.rename(columns={'SNO': 'Sl.', 'Price': 'price', 'Approval': 'approved'})
    if 'Note' not in csv_df.columns: csv_df['Note'] = ""
    csv_columns = ['Sl.', 'Name', 'Phone Number', 'Address', 'Quantity', 'Product', 'price', 'approved', 'Note', 'Date', 'Time']
    for c in csv_columns:
        if c not in csv_df.columns: csv_df[c] = ""
    csv_df = csv_df[csv_columns]
    
    export_columns = ["SNO", "Name", "Phone Number", "Address", "Quantity", "Product", "Price", "Approval", "Note", "Date", "Time"]
    for col in export_columns:
        if col not in export_df.columns: export_df[col] = ""
    export_df = export_df[export_columns]
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name="Orders")
        workbook = writer.book
        worksheet = writer.sheets['Orders']
        
        status_dv = DataValidation(type="list", formula1='"Pending,OK,Canceled,Talked,Not Picked,Sent to Pathao,Error,Incomplete"', allow_blank=True)
        worksheet.add_data_validation(status_dv)
        status_dv.add('H2:H10000') 
        
        pd.DataFrame({"Date": [sheet_date], "Total": [len(export_df)]}).to_excel(writer, index=False, sheet_name="Summary")
        for idx, prod in enumerate(product_list, start=1): writer.sheets['Summary'].cell(row=idx, column=5, value=prod)
        
        prod_dv = DataValidation(type="list", formula1=f"Summary!$E$1:$E${len(product_list)}", allow_blank=True)
        worksheet.add_data_validation(prod_dv)
        prod_dv.add('F2:F10000') 
        
        header_fill = PatternFill(start_color="e6f2ff", end_color="e6f2ff", fill_type="solid")
        sno_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        green_row_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid") 
        
        for cell in worksheet[1]: cell.fill, cell.font, cell.alignment, cell.border = header_fill, Font(bold=True, color="000000"), Alignment(horizontal="center", vertical="center"), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            prod_val = str(row[5].value).strip().lower()
            status_val = str(row[7].value).strip()
            note_val = str(row[8].value).strip()
            
            is_sent = (status_val == "Sent to Pathao") or ("Pathao ID" in note_val and status_val != "Incomplete")
            is_not_blender = (prod_val != "electric blender" and prod_val != "silver crest electric blender")
            is_incomplete = (status_val == "Incomplete")
            
            for cell in row:
                cell.border, cell.alignment = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')), Alignment(vertical="center")
                if cell.column == 1: cell.fill, cell.font, cell.alignment = sno_fill, Font(bold=True, color="FFFFFF"), Alignment(horizontal="center", vertical="center")
                elif is_incomplete: cell.fill, cell.font = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), Font(color="9C5700", bold=True)
                elif is_sent: cell.fill, cell.font = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), Font(color="9C0006")
                elif is_not_blender: cell.fill = green_row_fill

        worksheet.conditional_formatting.add('H2:H10000', CellIsRule(operator='equal', formula=['"OK"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"), font=Font(color="006100")))
        worksheet.conditional_formatting.add('H2:H10000', CellIsRule(operator='equal', formula=['"Pending"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), font=Font(color="9C5700")))
        worksheet.conditional_formatting.add('H2:H10000', CellIsRule(operator='equal', formula=['"Sent to Pathao"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), font=Font(color="9C0006")))
        worksheet.conditional_formatting.add('H2:H10000', CellIsRule(operator='equal', formula=['"Canceled"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), font=Font(color="9C0006")))
        worksheet.conditional_formatting.add('H2:H10000', CellIsRule(operator='equal', formula=['"Incomplete"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), font=Font(color="9C5700")))
        
        for col in worksheet.columns:
            max_len = 0
            for cell in col:
                try: max_len = max(max_len, len(str(cell.value)))
                except: pass
            worksheet.column_dimensions[col[0].column_letter].width = max_len + 2

    return output.getvalue(), csv_df.to_csv(index=False).encode('utf-8')


# --- APP LAYOUT HEADER ---
st.markdown("<br>", unsafe_allow_html=True)
col_logo, col_title, col_logout = st.columns([1.5, 6, 2])
with col_logo:
    if os.path.exists("logo.png"): st.image("logo.png", width=110)
with col_title: 
    st.markdown("<h2 class='main-header-title'>NWOP Dashboard</h2>", unsafe_allow_html=True)
    st.markdown("<div class='welcome-text'>Welcome back, Nazrul! Manage your orders seamlessly.</div>", unsafe_allow_html=True)
with col_logout: 
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🚪 Logout", type="secondary"):
        st.session_state.logged_in = False
        st.rerun()

# --- TABS ---
tab_workspace, tab_bulk_pathao, tab_ai_assistant, tab_merge, tab_history, tab_settings, tab_about = st.tabs(["🚀 Workspace", "🚚 Bulk Pathao", "🤖 AI Assistant", "🗂️ Merge", "📜 History", "⚙️ Settings", "ℹ️ About"])

with tab_workspace:
    st.sidebar.header("🛠️ Working Mode")
    mode = st.sidebar.radio("Select Input Mode:", ["Upload Chat History", "Live Scraping (Beta)"])
    
    st.sidebar.markdown("---")
    st.sidebar.header("🤖 AI Analyze Engine")
    analyze_engine_options = ["Groq: Llama 3.1 8B (Fast & Reliable)", "Groq: Llama 3.3 70B (High Logic)", "Google: Gemini 1.5 Flash (Vision)"]
    if 'analyze_engine' not in st.session_state: st.session_state.analyze_engine = analyze_engine_options[0]
    st.session_state.analyze_engine = st.sidebar.selectbox("Select Model for single AI Analyze:", analyze_engine_options, index=analyze_engine_options.index(st.session_state.analyze_engine))
    
    st.sidebar.markdown("---")
    st.sidebar.success(f"⏱️ **Last Extraction Checkpoint:**\n\n`{st.session_state.last_checkpoint}`\n\n*(Copy & use this as your next Start Time)*")

    if mode == "Upload Chat History":
        st.sidebar.header("📅 Extraction Filters")
        filter_type = st.sidebar.radio("Extract Data By:", ["All Time", "Specific Date", "Time Range (Copy-Paste)"])
        target_date_str, start_str, end_str = "", "", ""

        now = datetime.now(BD_TZ)
        safe_date_format = f"{now.month}/{now.day}/{now.strftime('%y')}"

        if filter_type == "Specific Date": target_date_str = st.sidebar.text_input("Enter Exact Date (e.g. 3/4/26):", safe_date_format)
        elif filter_type == "Time Range (Copy-Paste)":
            start_str = st.sidebar.text_input("Start Time:", st.session_state.last_checkpoint if st.session_state.last_checkpoint != "No record yet" else f"[{safe_date_format}, 9:21:30 PM]")
            end_str = st.sidebar.text_input("End Time:", f"[{safe_date_format}, 10:08:27 PM]")

        uploaded_file = st.file_uploader("📂 Upload WhatsApp Chat (.txt)", type="txt")

        if uploaded_file:
            if st.button("▶️ Start Regex Extraction", type="primary", use_container_width=True):
                with st.spinner("Processing file with Base Regex Core..."):
                    try:
                        content = uploaded_file.read().decode("utf-8").replace('\u200e', '').replace('\u200f', '')
                        lines = content.split('\n')
                        
                        messages, current_msg = [], None
                        for line in lines:
                            match = re.match(r'^\[(.*?),\s(.*?)\]\s.*?:', line)
                            if match:
                                if current_msg: messages.append(current_msg)
                                date_str, time_str = match.group(1), match.group(2)
                                msg_dt = get_datetime_obj(date_str, time_str)
                                dt_obj = msg_dt.date() if msg_dt else datetime.now(BD_TZ).date()
                                current_msg = {"date_obj": dt_obj, "date_str": date_str, "time_str": time_str, "msg_dt": msg_dt, "text": line}
                            else:
                                if current_msg: current_msg["text"] += "\n" + line
                        if current_msg: messages.append(current_msg)
                            
                        filtered_messages = []
                        start_dt = parse_copy_paste_time(start_str) if filter_type == "Time Range (Copy-Paste)" else None
                        end_dt = parse_copy_paste_time(end_str) if filter_type == "Time Range (Copy-Paste)" else None
                        
                        for msg in messages:
                            if filter_type == "All Time": filtered_messages.append(msg)
                            elif filter_type == "Specific Date" and target_date_str:
                                if msg["date_str"] == target_date_str.strip(): filtered_messages.append(msg)
                            elif filter_type == "Time Range (Copy-Paste)" and start_dt and end_dt and msg["msg_dt"]:
                                if start_dt <= msg["msg_dt"] <= end_dt: filtered_messages.append(msg)
                        
                        st.session_state.total_scanned = len(filtered_messages)
                        temp_orders, temp_ignored = [], []
                        phone_counts = {}
                        
                        for msg in filtered_messages:
                            data = extract_order_details(msg)
                            if data:
                                if data["status"] == "valid" and data["Address"] != "N/A" and len(data["Address"].strip()) > 3:
                                    del data["status"]
                                    ph = data['Phone Number']
                                    if ph != "N/A":
                                        phone_counts[ph] = phone_counts.get(ph, 0) + 1
                                        if phone_counts[ph] > 1: data["is_duplicate"] = True
                                    temp_orders.append(data)
                                else:
                                    if "status" in data: del data["status"]
                                    temp_ignored.append(data)
                        
                        st.session_state.ignored_messages = temp_ignored
                        if temp_orders:
                            st.session_state.all_orders = temp_orders 
                            st.session_state.sheet_date = "Time_Range_Export" if filter_type == "Time Range (Copy-Paste)" else target_date_str.replace('/', '-') if filter_type == "Specific Date" else f"Bulk_{datetime.now(BD_TZ).strftime('%d-%m-%y')}"
                            st.session_state.total_extracted_today += len(temp_orders)
                            
                            last_order = temp_orders[-1]
                            st.session_state.last_checkpoint = f"[{last_order['Date']}, {last_order['Time']}]"
                            
                            hist_src = f"Time Range ({start_str} to {end_str})" if filter_type == "Time Range (Copy-Paste)" else f"Date ({target_date_str})" if filter_type == "Specific Date" else "All Time"
                            log_task(f"Extracted {len(temp_orders)} orders via Regex. Source: {hist_src}.")
                            sync_profile_to_db()
                            st.success(f"Success! Found {len(temp_orders)} valid orders.")
                        else:
                            st.session_state.all_orders = []
                            st.error("No valid orders found with these filters.")
                    except Exception as e:
                        st.error(f"Error reading file: {e}")

    elif mode == "Live Scraping (Beta)":
        st.header("🔴 Auto-Scroll Live Scraper")
        if not SELENIUM_AVAILABLE:
            st.error("⚠️ Selenium is not installed! Please install via terminal: `pip install selenium webdriver-manager`")
        else:
            target_group = st.text_input("🎯 Live Target Group Name:", "ORDER COLLECTION")
            now = datetime.now(BD_TZ)
            safe_date_format = f"{now.month}/{now.day}/{now.strftime('%y')}"
            start_time_str = st.text_input("⏱️ Scrape From Exact Time (Copy-Paste):", st.session_state.last_checkpoint if st.session_state.last_checkpoint != "No record yet" else f"[{safe_date_format}, 7:16:42 PM]")

            if st.button("🚀 Launch WhatsApp & Fetch Orders", type="primary", use_container_width=True):
                target_limit_dt = parse_copy_paste_time(start_time_str)
                if not target_limit_dt:
                    st.error("⚠️ Start Time-এর ফরম্যাট ভুল! ব্র্যাকেট সহ ঠিকমতো কপি-পেস্ট করো।")
                else:
                    with st.spinner(f"Initializing Bot... Target Time: {target_limit_dt}"):
                        try:
                            chrome_options = Options()
                            chrome_options.add_argument("--disable-gpu")
                            chrome_options.add_argument("--no-sandbox")
                            chrome_options.add_argument("--disable-dev-shm-usage")
                            
                            service = ChromeService(ChromeDriverManager().install())
                            driver = webdriver.Chrome(service=service, options=chrome_options)
                            driver.get("https://web.whatsapp.com")
                            
                            st.info("🕒 Please scan the QR code within 60 seconds!")
                            WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "side")))
                            time.sleep(4)

                            search_box = None
                            try: search_box = driver.execute_script('return document.querySelector("#side [contenteditable=\'true\']");')
                            except: pass
                            
                            if not search_box:
                                search_xpaths = ['//*[@id="side"]//*[@contenteditable="true"]', '//div[@title="Search input textbox"]', '//div[@data-tab="3"]']
                                for xpath in search_xpaths:
                                    try:
                                        search_box = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, xpath)))
                                        if search_box: break
                                    except: pass
                            
                            try: search_box.click()
                            except: driver.execute_script("arguments[0].click();", search_box)
                            time.sleep(1)
                            try:
                                search_box.send_keys(Keys.CONTROL + "a")
                                search_box.send_keys(Keys.DELETE)
                            except: pass 
                            search_box.send_keys(target_group)
                            time.sleep(3)
                            
                            group_clicked = False
                            group_xpaths = [f"//span[@title='{target_group}']", f"//span[contains(@title, '{target_group}')]", f"//div[@title='{target_group}']", f"//span[text()='{target_group}']"]
                            for xpath in group_xpaths:
                                try:
                                    elem = WebDriverWait(driver, 4).until(EC.presence_of_element_located((By.XPATH, xpath)))
                                    driver.execute_script("arguments[0].click();", elem)
                                    group_clicked = True
                                    break
                                except: pass
                                    
                            if not group_clicked: raise Exception(f"Could not find or click group: '{target_group}'.")
                            st.info("✅ Group Found! Starting AUTO-SCROLL Engine...")
                            time.sleep(3)

                            safe_target_dt = target_limit_dt.replace(second=0)
                            scroll_attempts, max_scrolls = 0, 100 
                            
                            while scroll_attempts < max_scrolls:
                                msg_elements = driver.find_elements(By.XPATH, "//div[@data-pre-plain-text]")
                                if not msg_elements: break
                                first_msg_pre_text = msg_elements[0].get_attribute("data-pre-plain-text")
                                dt_str_match = re.search(r'\[(.*?)\]', first_msg_pre_text)
                                
                                if dt_str_match:
                                    dt_str = dt_str_match.group(1)
                                    parts = [p.strip() for p in dt_str.split(',')]
                                    oldest_dt = get_datetime_obj(parts[0], parts[1]) if len(parts) >= 2 else None
                                    
                                    if oldest_dt and oldest_dt > safe_target_dt:
                                        driver.execute_script("arguments[0].scrollIntoView(true);", msg_elements[0])
                                        time.sleep(1.5) 
                                        scroll_attempts += 1
                                    else: break
                                else: break
                            
                            time.sleep(2)

                            msg_elements = driver.find_elements(By.XPATH, "//div[@data-pre-plain-text]")
                            filtered_messages = []
                            for el in msg_elements: 
                                try:
                                    pre_text = el.get_attribute("data-pre-plain-text")
                                    text_span = el.find_element(By.XPATH, ".//span[contains(@class, 'selectable-text')]")
                                    msg_text = text_span.text
                                    
                                    dt_str_match = re.search(r'\[(.*?)\]', pre_text)
                                    if dt_str_match:
                                        dt_str = dt_str_match.group(1)
                                        parts = [p.strip() for p in dt_str.split(',')]
                                        msg_dt = get_datetime_obj(parts[0], parts[1]) if len(parts) >= 2 else None
                                        if not msg_dt: continue
                                        if msg_dt >= safe_target_dt:
                                            dt_obj = msg_dt.date()
                                            sender = pre_text.replace(f"[{dt_str}]", "").strip()
                                            date_str = parts[0] if '/' in parts[0] else parts[1]
                                            time_str = parts[0] if ':' in parts[0] else parts[1]
                                            
                                            raw_msg_constructed = f"[{date_str}, {time_str}] {sender} {msg_text}"
                                            if not is_whatsapp_system_message(raw_msg_constructed):
                                                filtered_messages.append({"date_obj": dt_obj, "date_str": date_str, "time_str": time_str, "msg_dt": msg_dt, "text": raw_msg_constructed})
                                except: pass
                            driver.quit()
                            
                            st.session_state.total_scanned = len(filtered_messages)
                            temp_orders, temp_ignored = [], []
                            phone_counts = {}
                            
                            for msg in filtered_messages:
                                data = extract_order_details(msg)
                                if data:
                                    if data["status"] == "valid" and data["Address"] != "N/A" and len(data["Address"].strip()) > 3:
                                        del data["status"]
                                        ph = data['Phone Number']
                                        if ph != "N/A":
                                            phone_counts[ph] = phone_counts.get(ph, 0) + 1
                                            if phone_counts[ph] > 1: data["is_duplicate"] = True
                                        temp_orders.append(data)
                                    else:
                                        if "status" in data: del data["status"]
                                        temp_ignored.append(data)
                            
                            st.session_state.ignored_messages = temp_ignored
                            if temp_orders:
                                st.session_state.all_orders = temp_orders
                                st.session_state.sheet_date = f"Live_Scrape_{datetime.now(BD_TZ).strftime('%d-%m-%y_%H%M')}"
                                st.session_state.total_extracted_today += len(temp_orders)
                                
                                last_order = temp_orders[-1]
                                st.session_state.last_checkpoint = f"[{last_order['Date']}, {last_order['Time']}]"
                                
                                log_task(f"Scraped {len(temp_orders)} orders via Live Scraper.")
                                sync_profile_to_db()
                                st.balloons()
                            else:
                                st.session_state.all_orders = []
                                st.warning("No valid new orders found.")
                        except Exception as e:
                            st.error(f"❌ Scraping Failed! {e}")
                            try: driver.quit() 
                            except: pass

    # --- DASHBOARD UI ---
    if st.session_state.all_orders or st.session_state.ignored_messages:
        
        suspect_keywords = ['taka', 'tk', 'টাকা', '/-', 'pice', 'pcs', 'পিস', 'blender', 'grinder', 'order', 'অর্ডার', 'thana', 'zilla', 'জেলা', 'থানা', 'গ্রাম']
        suspected_msgs, system_junk = [], []
        
        for ig in st.session_state.ignored_messages:
            text_lower = str(ig.get('Text', ig.get('RawText', ''))).lower()
            if any(kw in text_lower for kw in suspect_keywords) and len(text_lower) > 15: suspected_msgs.append(ig)
            else: system_junk.append(ig)
        
        st.markdown("<br><h4 style='text-align:center; color:#E0E0E0;'>📊 Data Reconciliation Report</h4>", unsafe_allow_html=True)
        col_r1, col_r2, col_r3, col_r4 = st.columns(4)
        col_r1.metric("🔍 Total Scanned", st.session_state.total_scanned)
        col_r2.metric("📦 Valid Orders", len(st.session_state.all_orders))
        col_r3.metric("🚨 Suspects / Errored", len(suspected_msgs))
        col_r4.metric("🗑️ System/Junk", len(system_junk))
        st.markdown("---")

        if st.session_state.all_orders:
            df = pd.DataFrame(st.session_state.all_orders)
            doubtful_orders = []
            passed_checks = 0
            total_checks = len(st.session_state.all_orders) * 3
            
            valid_prices = df[df['Price'].astype(int) > 0]['Price'].astype(int)
            avg_price = valid_prices.mean() if not valid_prices.empty else 0
            
            for i, row in enumerate(st.session_state.all_orders):
                issues = []
                p_check = re.match(r'^01[3-9]\d{8}$', str(row.get('Phone Number','')))
                pr_val = int(row.get('Price', 0))
                pr_check = pr_val > 0
                q_check = int(row.get('Quantity', 1)) > 0
                
                if p_check: passed_checks += 1
                else: issues.append("Invalid Phone")
                if pr_check: passed_checks += 1
                else: issues.append("Missing Price")
                if q_check: passed_checks += 1
                else: issues.append("Invalid Quantity")
                
                if avg_price > 0 and pr_val > 0 and pr_val < (avg_price * 0.5):
                    issues.append(f"📉 Low Price Alert (Avg is ৳{int(avg_price)})")
                
                if int(row.get('Quantity',1)) > 3:
                    issues.append("⚠️ High Qty (>3)")
                
                if str(row.get('Name','')).strip() == "N/A" or not str(row.get('Name','')).strip(): issues.append("Missing Name")
                elif any(h in str(row.get('Name','')).lower() for h in ['বাড়ি', 'বাড়ি', 'থানা', 'জেলা', 'রোড', 'road', 'গ্রাম', 'house']):
                    issues.append("Name looks like Address")

                if str(row.get('Address','')).strip() == "N/A" or not str(row.get('Address','')).strip(): issues.append("Missing Address")
                if row.get('is_duplicate', False): issues.append("⚠️ Duplicate Data")
                
                if issues: doubtful_orders.append({"id": row.get('id'), "order": row, "issues": issues})
            
            accuracy_score = round((passed_checks / total_checks) * 100, 1) if total_checks > 0 else 0
            
            st.markdown("<br>", unsafe_allow_html=True)
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("📦 Orders", len(st.session_state.all_orders))
            m2.metric("💰 Revenue", f"৳ {sum(int(o.get('Price',0)) for o in st.session_state.all_orders)}")
            m3.metric("🎯 Accuracy", f"{accuracy_score}%")
            m4.metric("📈 Session Total", st.session_state.total_extracted_today)

            if doubtful_orders:
                st.error(f"⚠️ Action Required: Found {len(doubtful_orders)} doubtful or duplicate entries!")
                with st.expander("🚨 REVIEW DOUBTFUL ENTRIES", expanded=True):
                    for dob in doubtful_orders:
                        o_id = dob['id']
                        issue_text = ', '.join(dob['issues'])
                        st.markdown(f"""
                            <div class="doubt-card">
                                <div style="flex: 1;">
                                    <strong style="color: #FFC107;">Issue:</strong> {issue_text}<br>
                                    <span style="color: #E0E0E0; font-size:0.9rem;"><strong>Name:</strong> {dob['order'].get('Name')} | <strong>Phone:</strong> {dob['order'].get('Phone Number')}</span>
                                </div>
                                <div><a href="#order-{o_id}" class="fix-btn">Fix 🔨</a></div>
                            </div>
                        """, unsafe_allow_html=True)
            
            col_m1, col_m2 = st.columns([4, 1.5])
            with col_m1: st.markdown("### 📋 Manage Orders")
            with col_m2:
                img_file = st.file_uploader("📸 Scan Image to Order", type=["png", "jpg", "jpeg"])
                if img_file:
                    if st.button("✨ Extract from Image", type="primary"):
                        if not st.session_state.gemini_api_key or not GEMINI_AVAILABLE:
                            st.error("⚠️ Gemini API Key required for Vision. Check Settings.")
                        else:
                            with st.spinner("🤖 Vision AI is reading the image..."):
                                extracted_data = extract_from_image_vision(img_file, st.session_state.gemini_api_key)
                                if extracted_data:
                                    raw_ph = str(extracted_data.get('Phone Number', 'N/A'))
                                    final_ph = format_phone_number(raw_ph)
                                    
                                    new_vision_order = {
                                        "id": str(uuid.uuid4()), "Date": datetime.now(BD_TZ).strftime("%d/%m/%y"), "Time": datetime.now(BD_TZ).strftime("%I:%M %p"),
                                        "Name": extracted_data.get('Name', 'N/A'), "Phone Number": final_ph, "Address": extracted_data.get('Address', 'N/A'),
                                        "Product": extracted_data.get('Product', 'Electric Blender'), "Quantity": int(extracted_data.get('Quantity', 1)), "Price": int(extracted_data.get('Price', 0)),
                                        "Approval": "Pending", "Note": "Extracted via Vision AI", "is_duplicate": False, "RawText": "📸 Extracted from uploaded image.", "Method": "👁️ Vision AI",
                                    }
                                    learn_new_product(new_vision_order['Product'])
                                    st.session_state.all_orders.insert(0, new_vision_order)
                                    st.success("✅ Image processed successfully!")
                                    time.sleep(1)
                                    st.rerun()
                                else: st.error("❌ Failed to read text from image.")
                
                if st.button("➕ Add Manual Order", type="secondary"):
                    new_manual_order = {
                        "id": str(uuid.uuid4()), "Date": datetime.now(BD_TZ).strftime("%d/%m/%y"), "Time": datetime.now(BD_TZ).strftime("%I:%M %p"),
                        "Name": "", "Phone Number": "", "Address": "", "Product": "Electric Blender", "Quantity": 1, "Price": 0,
                        "Approval": "Pending", "Note": "Manual Entry", "is_duplicate": False, "RawText": "✍️ This order was added manually.", "Method": "✍️ Manual"
                    }
                    st.session_state.all_orders.insert(0, new_manual_order)
                    st.rerun()

            for i, row in enumerate(st.session_state.all_orders):
                if 'id' not in row: row['id'] = str(uuid.uuid4())
                o_id = row['id']
                
                dup_tag = " (⚠️ Duplicate)" if row.get('is_duplicate', False) else ""
                final_title = f"Order: {row.get('Name', 'N/A')} | ৳{row.get('Price', 0)} | 📞 {row.get('Phone Number', 'N/A')} | 🕒 {row.get('Time', '')} | {row.get('Method', '')}{dup_tag}"
                
                st.markdown(f'<div id="order-{o_id}" style="position: relative; top: -60px;"></div>', unsafe_allow_html=True)
                
                with st.expander(final_title, expanded=False):
                    clean_raw = row.get('RawText', 'N/A').replace('\n', '<br>')
                    clean_raw = re.split(r'^\[.*?\] .*?:\s', clean_raw, maxsplit=1)[-1] 
                    
                    st.markdown(f"""
                    <div style="display: flex; flex-direction: column; align-items: flex-start; margin-bottom: 20px;">
                        <div class='wa-bubble'>
                            <div style="font-size: 11px; opacity: 0.8; margin-bottom: 5px; font-weight: 600; text-transform: uppercase;">💬 Original Message</div>
                            <div style="font-size: 15px;">{clean_raw}</div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    c1, c2 = st.columns([1, 1])
                    with c1:
                        new_name = st.text_input("👤 Name:", value=row.get('Name',''), key=f"name_{o_id}")
                        new_addr = st.text_input("🏠 Address:", value=row.get('Address',''), key=f"addr_{o_id}")
                        new_phone = st.text_input("📱 Phone:", value=row.get('Phone Number',''), key=f"phone_{o_id}")
                        
                        p_val = row.get('Product', 'Electric Blender')
                        if p_val not in st.session_state.product_list: st.session_state.product_list.append(p_val)
                        idx_prod = st.session_state.product_list.index(p_val)
                        new_prod = st.selectbox("📦 Item:", st.session_state.product_list, index=idx_prod, key=f"prod_{o_id}")
                        
                        new_note = st.text_input("📝 Note:", value=row.get('Note', ''), key=f"note_{o_id}")
                        
                        st.session_state.all_orders[i]['Name'] = new_name
                        st.session_state.all_orders[i]['Address'] = new_addr
                        st.session_state.all_orders[i]['Phone Number'] = new_phone
                        st.session_state.all_orders[i]['Product'] = new_prod
                        st.session_state.all_orders[i]['Note'] = new_note
                        
                    with c2:
                        col_p, col_q = st.columns(2)
                        with col_p:
                            new_price = st.number_input("💰 Price (৳):", value=int(row.get('Price',0)), min_value=0, key=f"price_{o_id}")
                        with col_q:
                            new_qty = st.number_input("⚖️ Qty:", value=int(row.get('Quantity',1)), min_value=0, key=f"qty_{o_id}")
                            
                        st.session_state.all_orders[i]['Price'] = new_price
                        st.session_state.all_orders[i]['Quantity'] = new_qty
                        
                        status_list = ["Pending", "OK", "Canceled", "Talked", "Not Picked", "Sent to Pathao", "Error", "Incomplete"]
                        current_idx = status_list.index(row.get('Approval','Pending')) if row.get('Approval') in status_list else 0
                        new_status = st.selectbox("Status:", status_list, index=current_idx, key=f"status_{o_id}")
                        st.session_state.all_orders[i]['Approval'] = new_status
                        
                    col_rm1, col_rm2, col_rm3, col_rm4 = st.columns([2, 2, 2, 1])
                    with col_rm1:
                        st.markdown(f'''<a href="tel:{new_phone}" style="display:inline-block; text-align:center; width:100%; background: linear-gradient(135deg, #10B981, #059669); color:white; padding:10px 15px; border-radius:25px; margin-top:20px; font-weight:bold; box-shadow: 0px 4px 10px rgba(16, 185, 129, 0.3); text-decoration:none;">📞 Call Customer</a>''', unsafe_allow_html=True)
                    with col_rm2:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("🚚 Send to Pathao", key=f"pathao_btn_{o_id}", help="Push order directly to Pathao", type="primary"):
                            if not st.session_state.pathao_client_id or not st.session_state.pathao_email: st.error("⚠️ Credentials Missing! Go to Settings.")
                            elif not new_phone or new_phone == "N/A": st.error("⚠️ Valid Phone Number required.")
                            elif new_price == 0: st.error("⚠️ Price cannot be 0.")
                            else:
                                with st.spinner("Connecting to Pathao API..."):
                                    success_status, pathao_msg = send_to_pathao_api(
                                        st.session_state.all_orders[i], 
                                        st.session_state.pathao_client_id, 
                                        st.session_state.pathao_client_secret, 
                                        st.session_state.pathao_store_id,
                                        st.session_state.pathao_email,
                                        st.session_state.pathao_password
                                    )
                                    if success_status == True:
                                        st.session_state.all_orders[i]['Approval'] = "Sent to Pathao"
                                        st.session_state.all_orders[i]['Note'] = f"✅ Pathao ID: {pathao_msg}"
                                        st.session_state.all_orders[i]['id'] = str(uuid.uuid4()) 
                                        log_task(f"Pushed order {new_name} to Pathao. ID: {pathao_msg}")
                                        st.success(f"Success! Consignment ID: {pathao_msg}")
                                        time.sleep(1)
                                        st.rerun()
                                    elif success_status == "INCOMPLETE":
                                        st.session_state.all_orders[i]['Approval'] = "Incomplete"
                                        st.session_state.all_orders[i]['Note'] = f"⚠️ ID: {pathao_msg} (Dhaka, Banani)"
                                        st.session_state.all_orders[i]['id'] = str(uuid.uuid4()) 
                                        log_task(f"Pushed order {new_name} to Pathao (INCOMPLETE). ID: {pathao_msg}")
                                        st.warning(f"Created but INCOMPLETE: {pathao_msg}")
                                        time.sleep(2)
                                        st.rerun()
                                    else: st.error(pathao_msg)
                                        
                    with col_rm3:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("🤖 AI Analyze", key=f"ai_analyze_btn_{o_id}"):
                            eng = st.session_state.analyze_engine
                            use_groq = "Groq" in eng
                            use_gemini = "Gemini" in eng
                            
                            if use_groq and not st.session_state.groq_api_key: st.error("⚠️ Groq API not available.")
                            elif use_gemini and not st.session_state.gemini_api_key: st.error("⚠️ Gemini API not available.")
                            else:
                                with st.spinner("⚡ AI is fixing this message..."):
                                    new_data = analyze_single_order(row.get('RawText',''), eng, st.session_state.groq_api_key, st.session_state.gemini_api_key)

                                    if new_data:
                                        st.session_state.all_orders[i]['Name'] = str(new_data.get('Name', new_name)).strip()
                                        st.session_state.all_orders[i]['Address'] = str(new_data.get('Address', new_addr)).strip()
                                        
                                        ai_ph = format_phone_number(str(new_data.get('Phone Number', '')))
                                        st.session_state.all_orders[i]['Phone Number'] = ai_ph if ai_ph != "N/A" else new_phone
                                        
                                        p_val = str(new_data.get('Product', new_prod)).strip()
                                        learn_new_product(p_val)
                                        st.session_state.all_orders[i]['Product'] = p_val
                                        
                                        st.session_state.all_orders[i]['Price'] = int(new_data.get('Price', new_price))
                                        st.session_state.all_orders[i]['Quantity'] = int(new_data.get('Quantity', new_qty))
                                        st.session_state.all_orders[i]['Method'] = "🤖 AI Corrected"
                                        
                                        st.session_state.all_orders[i]['id'] = str(uuid.uuid4())
                                        
                                        st.success("✅ AI perfectly fixed the data!")
                                        time.sleep(1)
                                        st.rerun()
                                    else:
                                        st.error("❌ AI couldn't parse. Try manual entry.")
                                
                    with col_rm4:
                        st.markdown("<br>", unsafe_allow_html=True)
                        if st.button("🗑️", key=f"del_btn_{o_id}", help="Remove Order"):
                            st.session_state.all_orders = [o for o in st.session_state.all_orders if o['id'] != o_id]
                            st.rerun()

            st.markdown("---")
            filename = f"NWOP_Orders_{st.session_state.sheet_date}.xlsx"
            excel_bytes, csv_bytes = generate_excel_bytes(st.session_state.all_orders, st.session_state.sheet_date, st.session_state.product_list)
            
            col_d1, col_d2 = st.columns(2)
            with col_d1:
                st.download_button(label="📥 Download Excel File", data=excel_bytes, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
            with col_d2:
                csv_filename = f"NWOP_Orders_{st.session_state.sheet_date}.csv"
                st.download_button(label="📊 Download CSV (For Google Sheets)", data=csv_bytes, file_name=csv_filename, mime="text/csv", type="secondary", use_container_width=True)

        if suspected_msgs:
            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander(f"🚨 SUSPECTED MISSED ORDERS ({len(suspected_msgs)} items)", expanded=True):
                for idx, sm in enumerate(suspected_msgs):
                    sm_id = sm.get("id", f"susp_{idx}")
                    st.caption(f"🕒 {sm.get('Date', '')} - {sm.get('Time', '')}")
                    raw_txt = sm.get('Text', sm.get('RawText', ''))
                    clean_display = re.split(r'^\[.*?\] .*?:\s', raw_txt, maxsplit=1)[-1]
                    
                    st.error(f"**Reason:** {sm.get('Reason', 'Unknown Keyword Flag')} \n\n**Original Message:**\n{clean_display}")
                    
                    if st.button("➕ List as Order", key=f"add_susp_{sm_id}"):
                        new_order = {
                            "id": str(uuid.uuid4()), "Date": sm.get("Date", ""), "Time": sm.get("Time", ""),
                            "Name": "N/A", "Phone Number": "N/A", "Address": "N/A", "Product": "Electric Blender", "Quantity": 1, "Price": 0,
                            "Approval": "Pending", "Note": "From Suspected List", "is_duplicate": False,
                            "RawText": raw_txt, "Method": "✍️ Manual",
                        }
                        st.session_state.all_orders.insert(0, new_order)
                        st.session_state.ignored_messages = [m for m in st.session_state.ignored_messages if m.get("id") != sm_id]
                        st.rerun()
                    
        if system_junk:
            st.markdown("<br>", unsafe_allow_html=True)
            with st.expander(f"🗑️ System Messages / Junk ({len(system_junk)} items)", expanded=False):
                st.info("Egulo normal text/system message. Ete kono order info nai.")
                for idx, jm in enumerate(system_junk):
                    jm_id = jm.get("id", f"junk_{idx}")
                    st.caption(f"🕒 {jm.get('Date', '')} - {jm.get('Time', '')}")
                    raw_txt = jm.get('Text', jm.get('RawText', ''))
                    clean_display = re.split(r'^\[.*?\] .*?:\s', raw_txt, maxsplit=1)[-1]
                    
                    st.warning(f"**Reason:** {jm.get('Reason', 'System Junk Filtered')} \n\n**Original Message:**\n{clean_display}")
                    
                    if st.button("➕ List as Order", key=f"add_junk_{jm_id}"):
                        new_order = {
                            "id": str(uuid.uuid4()), "Date": jm.get("Date", ""), "Time": jm.get("Time", ""),
                            "Name": "N/A", "Phone Number": "N/A", "Address": "N/A", "Product": "Electric Blender", "Quantity": 1, "Price": 0,
                            "Approval": "Pending", "Note": "From Junk List", "is_duplicate": False,
                            "RawText": raw_txt, "Method": "✍️ Manual",
                        }
                        st.session_state.all_orders.insert(0, new_order)
                        st.session_state.ignored_messages = [m for m in st.session_state.ignored_messages if m.get("id") != jm_id]
                        st.rerun()

# 🌟 BULK PATHAO UPLOAD TAB 🌟
with tab_bulk_pathao:
    st.header("🚚 Bulk Pathao Upload")
    st.info("Upload your exported NWOP Excel/CSV file to send multiple orders to Pathao at once. You can edit before sending and stop anytime!")
    
    bulk_file = st.file_uploader("📂 Upload NWOP Excel/CSV", type=["xlsx", "csv"], key="bulk_uploader")
    
    if bulk_file:
        file_hash = hashlib.md5(bulk_file.getvalue()).hexdigest()
        
        if st.session_state.get('bulk_file_hash') != file_hash:
            st.session_state.bulk_file_hash = file_hash
            if bulk_file.name.endswith('.csv'): b_df = pd.read_csv(bulk_file, dtype=str)
            else: b_df = pd.read_excel(bulk_file, sheet_name="Orders", dtype=str)
            
            b_df.dropna(how='all', inplace=True)
            b_df = b_df.rename(columns={'Sl.': 'SNO', 'price': 'Price', 'approved': 'Approval'})
            
            bulk_orders = []
            for _, row in b_df.iterrows():
                if pd.isna(row.get('Phone Number')) and pd.isna(row.get('Name')): continue
                
                appr = str(row.get('Approval', 'Pending'))
                note = str(row.get('Note', ''))
                is_sent = "Sent to Pathao" in appr or ("Pathao ID" in note and "Incomplete" not in appr)
                
                raw_ph = str(row.get('Phone Number', 'N/A')).replace('.0', '').strip()
                if raw_ph.lower() == 'nan': raw_ph = 'N/A'
                clean_ph = format_phone_number(raw_ph) if raw_ph != "N/A" else "N/A"
                
                raw_qty = str(row.get('Quantity', 1)).replace('N/A', '1').strip().lower()
                if raw_qty == 'nan' or not raw_qty: raw_qty = '1'
                try: final_qty = int(float(raw_qty))
                except: final_qty = 1
                
                raw_price = str(row.get('Price', 0)).replace('N/A', '0').strip().lower()
                if raw_price == 'nan' or not raw_price: raw_price = '0'
                try: final_price = int(float(raw_price))
                except: final_price = 0
                
                raw_name = str(row.get('Name', 'N/A'))
                if raw_name.lower() == 'nan': raw_name = 'N/A'
                
                raw_addr = str(row.get('Address', 'N/A'))
                if raw_addr.lower() == 'nan': raw_addr = 'N/A'
                
                raw_prod = str(row.get('Product', 'Electric Blender'))
                if raw_prod.lower() == 'nan': raw_prod = 'Electric Blender'
                
                bulk_orders.append({
                    "id": str(uuid.uuid4()), "Name": raw_name,
                    "Phone Number": clean_ph,
                    "Address": raw_addr, "Product": raw_prod,
                    "Quantity": final_qty,
                    "Price": final_price,
                    "Approval": appr, "Note": note, "is_sent": is_sent,
                    "Date": str(row.get('Date', datetime.now(BD_TZ).strftime("%d/%m/%y"))),
                    "Time": str(row.get('Time', datetime.now(BD_TZ).strftime("%I:%M %p"))),
                })
                
            st.session_state.bulk_orders = bulk_orders
            st.session_state.bulk_processed_orders = None
            st.session_state.is_sending_bulk = False
            st.session_state.bulk_sending_list = []
            st.session_state.bulk_sent_list = []
            st.session_state.bulk_results = []
            
    if st.session_state.get('bulk_orders') and not st.session_state.get('bulk_processed_orders') and not st.session_state.get('is_sending_bulk'):
        pending_orders = [o for o in st.session_state.bulk_orders if not o['is_sent']]
        sent_already = [o for o in st.session_state.bulk_orders if o['is_sent']]
        
        st.write(f"📊 **Total Orders:** {len(st.session_state.bulk_orders)} | ⏳ **Pending:** {len(pending_orders)} | ✅ **Already Sent:** {len(sent_already)}")
        
        if pending_orders:
            st.write("### 📝 Review & Edit Pending Orders")
            st.info("You can edit the Name, Phone, Address, Price, and Note directly below. **To skip/remove an order, check the '❌ Drop' box.**")
            
            # 🌟 ADDED '❌ Drop' COLUMN FOR EASY ROW REMOVAL 🌟
            pending_df = pd.DataFrame(pending_orders)[["id", "Name", "Phone Number", "Address", "Price", "Quantity", "Note"]]
            pending_df.insert(0, '❌ Drop', False)
            
            edited_df = st.data_editor(pending_df, hide_index=True, use_container_width=True, disabled=["id"])
            edited_records = edited_df.to_dict('records')
            
            col1, col2 = st.columns([1, 4])
            with col1:
                start_send = st.button("🚀 Confirm & Send to Pathao", type="primary", use_container_width=True)
                
            if start_send:
                if not st.session_state.pathao_client_id or not st.session_state.pathao_email:
                    st.error("⚠️ Credentials Missing! Go to Settings.")
                else:
                    st.session_state.bulk_results = []
                    final_pending = []
                    ed_dict = {str(row['id']): row for row in edited_records}
                    
                    for orig in pending_orders:
                        o_id = str(orig['id'])
                        if o_id in ed_dict:
                            ed = ed_dict[o_id]
                            # Check if the user dropped the row
                            if ed.get('❌ Drop', False):
                                orig['Approval'] = "Not Picked"
                                orig['Note'] = "❌ Skipped by User"
                                st.session_state.bulk_results.append(orig)
                                continue
                                
                            orig['Name'] = ed.get('Name', orig['Name'])
                            orig['Phone Number'] = ed.get('Phone Number', orig['Phone Number'])
                            orig['Address'] = ed.get('Address', orig['Address'])
                            try: orig['Price'] = int(float(ed.get('Price', orig['Price'])))
                            except: orig['Price'] = 0
                            try: orig['Quantity'] = int(float(ed.get('Quantity', orig['Quantity'])))
                            except: orig['Quantity'] = 1
                            orig['Note'] = ed.get('Note', orig['Note'])
                            
                            final_pending.append(orig)
                    
                    st.session_state.bulk_sending_list = final_pending
                    st.session_state.bulk_sent_list = sent_already
                    st.session_state.is_sending_bulk = True
                    st.session_state.bulk_total_to_send = len(final_pending)
                    st.rerun()

    if st.session_state.get('is_sending_bulk'):
        total = st.session_state.bulk_total_to_send
        done = len(st.session_state.bulk_results) - len([o for o in st.session_state.bulk_results if o.get('Note') == "❌ Skipped by User"])
        
        st.warning("⚠️ **Sending in progress... Do not refresh the page.**")
        st.progress(done / total if total > 0 else 1.0, text=f"Processing order {done+1} of {total}...")
        
        col1, col2 = st.columns([1, 5])
        with col1:
            if st.button("🛑 Cancel / Stop Sending", type="secondary", use_container_width=True):
                st.session_state.is_sending_bulk = False
                for o in st.session_state.bulk_sending_list:
                    o['Approval'] = "Canceled"
                    o['Note'] = "❌ Canceled by user"
                    st.session_state.bulk_results.append(o)
                st.session_state.bulk_sending_list = []
                st.session_state.bulk_processed_orders = st.session_state.bulk_sent_list + st.session_state.bulk_results
                st.rerun()
                
        if st.session_state.is_sending_bulk:
            if len(st.session_state.bulk_sending_list) > 0:
                order = st.session_state.bulk_sending_list.pop(0)
                
                if order['Phone Number'] == "N/A" or int(order['Price']) == 0:
                    order['Approval'] = "Error"
                    order['Note'] = "❌ Missing Phone or Price"
                else:
                    success_status, msg = send_to_pathao_api(order, st.session_state.pathao_client_id, st.session_state.pathao_client_secret, st.session_state.pathao_store_id, st.session_state.pathao_email, st.session_state.pathao_password)
                    if success_status == True:
                        order['Approval'] = "Sent to Pathao"
                        order['Note'] = f"✅ Pathao ID: {msg}"
                        order['is_sent'] = True
                    elif success_status == "INCOMPLETE":
                        order['Approval'] = "Incomplete"
                        order['Note'] = f"⚠️ ID: {msg} (Dhaka, Banani)"
                        order['is_sent'] = True 
                    else:
                        order['Approval'] = "Error"
                        order['Note'] = f"❌ {msg}"
                
                st.session_state.bulk_results.append(order)
                time.sleep(0.5) 
                st.rerun()
            else:
                st.session_state.is_sending_bulk = False
                st.session_state.bulk_processed_orders = st.session_state.bulk_sent_list + st.session_state.bulk_results
                st.success("✅ All pending orders processed!")
                st.rerun()

    if st.session_state.get('bulk_processed_orders') and not st.session_state.get('is_sending_bulk'):
        st.write("### 📝 Results")
        res_df = pd.DataFrame(st.session_state.bulk_processed_orders)[["Name", "Phone Number", "Approval", "Note"]]
        st.dataframe(res_df)
        
        bulk_filename = f"NWOP_Bulk_Result_{datetime.now(BD_TZ).strftime('%d-%m-%y_%H%M')}.xlsx"
        bulk_excel, bulk_csv = generate_excel_bytes(st.session_state.bulk_processed_orders, datetime.now(BD_TZ).strftime('%d-%m-%y'), st.session_state.product_list)
        
        col_b1, col_b2 = st.columns(2)
        with col_b1:
            st.download_button(label="📥 Download Updated Excel (Red Marked)", data=bulk_excel, file_name=bulk_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
        with col_b2:
            st.download_button(label="📊 Download Updated CSV", data=bulk_csv, file_name=bulk_filename.replace('.xlsx', '.csv'), mime="text/csv", type="secondary", use_container_width=True)

# 🌟 NATIVE CHAT UI AI ASSISTANT TAB 🌟
with tab_ai_assistant:
    st.header("🤖 NWOP AI Assistant")
    st.caption("Powered by Groq & Google Gemini")
    
    col_c1, col_c2 = st.columns([5, 1])
    with col_c2:
        if st.button("🧹 Clear Chat"):
            st.session_state.chat_history = [{"role": "assistant", "content": "হ্যালো! আমি NWOP এআই। আপনার বিজনেস ডাটা বা অন্য যেকোনো বিষয়ে আমি সাহায্য করতে প্রস্তুত।"}]
            st.rerun()
    
    if not GROQ_AVAILABLE or not st.session_state.groq_api_key:
        st.warning("⚠️ For the best chat experience, set your Groq API Key in Settings.")
        
    total_orders = len(st.session_state.all_orders)
    total_revenue = sum(int(o.get('Price', 0)) for o in st.session_state.all_orders)
    
    db_history_context = "No recent database history found."
    if SUPABASE_AVAILABLE and st.session_state.get('supabase'):
        try:
            orders_res = st.session_state.supabase.table('nwop_orders').select('order_data').eq('user_id', st.session_state.user.id).order('created_at', desc=True).limit(50).execute()
            if orders_res.data:
                db_orders = [o['order_data'] for o in orders_res.data]
                db_history_context = json.dumps(db_orders, ensure_ascii=False)[:2500] 
        except: pass
    
    for msg in st.session_state.chat_history:
        with st.chat_message(msg["role"]): st.markdown(msg["content"])
            
    if prompt := st.chat_input("Ask me anything about your business or general topics..."):
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        with st.chat_message("user"): st.markdown(prompt)

        with st.chat_message("assistant"):
            message_placeholder = st.empty()
            with st.spinner("🤖 Thinking..."):
                try:
                    system_prompt = f"""
                    You are "NWOP AI" (Nazrul's Whatsapp Order Parser AI), a highly intelligent, friendly, and professional conversational assistant.
                    Your creator/developer is Nazrul Rana. 
                    Developer Details:
                    - Name: Nazrul Rana
                    - Education: B.Sc. in Computer Science and Engineering (CSE) from Southeast University, Bangladesh.
                    - WhatsApp: 01641434000
                    - Facebook: https://www.facebook.com/nazrulranaxD.s
                    - LinkedIn: https://www.linkedin.com/in/nazrulhuda/
                    
                    Your Capabilities:
                    - Process, analyze, and discuss the user's business data. 
                    - CURRENT SESSION CONTEXT: Orders Today: {total_orders}, Revenue: {total_revenue} Tk. Products: {', '.join(st.session_state.product_list)}
                    - RECENT DATABASE HISTORY CONTEXT: {db_history_context}
                    - Excellent at casual chatting (Adda), telling jokes, current weather, Bangladesh news.
                    """
                    
                    if "Groq" in st.session_state.analyze_engine and st.session_state.groq_api_key:
                        client = Groq(api_key=st.session_state.groq_api_key)
                        model_use = "llama-3.3-70b-versatile" if "70B" in st.session_state.analyze_engine else "llama-3.1-8b-instant"
                        api_messages = [{"role": "system", "content": system_prompt}] + [{"role": m["role"], "content": m["content"]} for m in st.session_state.chat_history]
                        response = client.chat.completions.create(messages=api_messages, model=model_use, temperature=0.6)
                        ai_reply = response.choices[0].message.content
                    elif st.session_state.gemini_api_key:
                        genai.configure(api_key=st.session_state.gemini_api_key)
                        model = genai.GenerativeModel('gemini-1.5-flash-latest')
                        chat_history_str = "\n".join([f"{m['role']}: {m['content']}" for m in st.session_state.chat_history])
                        full_prompt = f"{system_prompt}\n\nChat History:\n{chat_history_str}"
                        response = model.generate_content(full_prompt)
                        ai_reply = response.text.strip()
                    else:
                        ai_reply = "⚠️ Please set your Groq or Gemini API Key in Settings to use the AI Assistant."
                        
                    message_placeholder.markdown(ai_reply)
                    st.session_state.chat_history.append({"role": "assistant", "content": ai_reply})
                except Exception as e:
                    message_placeholder.error(f"AI Error: {str(e)}")

with tab_merge:
    st.header("🗂️ Excel Merger (Smart Sorter)")
    st.info("Here you can upload multiple NWOP Excel files. The app will merge them, remove duplicates, sort them perfectly by Date & Time, and create a single Master File without losing phone number leading '0's!")
    
    merge_mode = st.radio("Select Merge Action:", ["Standard Order Merge", "Bulk Pathao Master Merge"], horizontal=True)
    
    uploaded_excels = st.file_uploader("📂 Select multiple Excel files", type=["xlsx"], accept_multiple_files=True)
    
    if uploaded_excels and len(uploaded_excels) > 0:
        with st.spinner("Processing files..."):
            try:
                all_dfs = []
                for file in uploaded_excels:
                    df = pd.read_excel(file, sheet_name="Orders", dtype=str)
                    if 'Phone Number' in df.columns:
                        df['Phone Number'] = df['Phone Number'].fillna("N/A").apply(lambda x: format_phone_number(str(x).replace('.0', '')))
                    all_dfs.append(df)
                
                merged_df = pd.concat(all_dfs, ignore_index=True)
                
                merged_df['sort_dt'] = merged_df.apply(lambda r: get_datetime_obj(str(r.get('Date', '')), str(r.get('Time', ''))) or datetime.min, axis=1)
                merged_df = merged_df.sort_values(by='sort_dt')
                merged_df = merged_df.drop_duplicates(subset=['Phone Number'], keep='last')
                merged_df = merged_df.drop(columns=['sort_dt'])
                
                merged_df['Quantity'] = pd.to_numeric(merged_df['Quantity'], errors='coerce').fillna(1).astype(int)
                merged_df['is_multi'] = merged_df['Quantity'] > 1
                merged_df = merged_df.sort_values(by=['is_multi']).drop(columns=['is_multi'])
                
                merged_df['SNO'] = range(1, len(merged_df) + 1)
                
                export_columns = ["SNO", "Name", "Phone Number", "Address", "Quantity", "Product", "Price", "Approval", "Note", "Date", "Time"]
                excel_df = merged_df.copy()
                for c in export_columns:
                    if c not in excel_df.columns: excel_df[c] = ""
                excel_df = excel_df[export_columns]
                
                merged_orders = excel_df.to_dict('records')
                bulk_excel, bulk_csv = generate_excel_bytes(merged_orders, datetime.now(BD_TZ).strftime('%d-%m-%y'), st.session_state.product_list)
                
                st.success(f"✅ Successfully combined {len(uploaded_excels)} files into {len(excel_df)} unique sorted orders!")
                
                col_md1, col_md2 = st.columns(2)
                with col_md1:
                    st.download_button(label="📥 Download Master Excel", data=bulk_excel, file_name=f"NWOP_Master_{datetime.now(BD_TZ).strftime('%d-%m-%y')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary", use_container_width=True)
                with col_md2:
                    st.download_button(label="📊 Download CSV (For Google Sheets)", data=bulk_csv, file_name=f"NWOP_Master_{datetime.now(BD_TZ).strftime('%d-%m-%y')}.csv", mime="text/csv", type="secondary", use_container_width=True)
            except Exception as e:
                st.error(f"Error processing files: {e}. Please make sure you uploaded valid NWOP Excel files.")

with tab_history:
    st.header("📜 Task History & Logs")
    if not st.session_state.task_history:
        st.info("No activity recorded yet.")
    else:
        for task in st.session_state.task_history: st.markdown(task, unsafe_allow_html=True)
        if st.button("Clear History", type="secondary"):
            st.session_state.task_history = []
            st.session_state.last_checkpoint = "No record yet"
            sync_profile_to_db()
            st.rerun()

with tab_settings:
    st.header("⚙️ NWOP Settings")
    st.markdown("**Version:** NWOP v52.1 (The Final UI & Sync Fix)")
    
    st.markdown("### ⚡ AI Engine Keys")
    new_api_key = st.text_input("Groq API Key (Speed & High Quota)", type="password", value=st.session_state.groq_api_key, placeholder="gsk_...")
    new_gem_key = st.text_input("Gemini API Key (High Quota & Vision)", type="password", value=st.session_state.gemini_api_key, placeholder="AIzaSy...")
    
    st.markdown("---")
    st.markdown("### 🚚 Pathao Courier API Credentials")
    st.info("Pathao uses Email and Password for authenticating Merchant Orders. Please provide your Pathao Panel login credentials.")
    new_p_client = st.text_input("Pathao Client ID", value=st.session_state.pathao_client_id, placeholder="Client ID from Pathao Panel")
    new_p_secret = st.text_input("Pathao Client Secret", type="password", value=st.session_state.pathao_client_secret, placeholder="Client Secret")
    new_p_email = st.text_input("Pathao Email (Username)", value=st.session_state.pathao_email, placeholder="your_email@domain.com")
    new_p_pass = st.text_input("Pathao Password", type="password", value=st.session_state.pathao_password, placeholder="Your Pathao Login Password")
    new_p_store = st.text_input("Pathao Store ID", value=st.session_state.pathao_store_id, placeholder="e.g. 12345")
    
    st.markdown("---")
    st.markdown("### 📦 Custom Products")
    new_custom_prod = st.text_input("Add Custom Product to List permanently:", placeholder="e.g. Digital Scale")
    
    if st.button("Save All Settings", type="primary"):
        st.session_state.groq_api_key = new_api_key
        st.session_state.gemini_api_key = new_gem_key
        st.session_state.pathao_client_id = new_p_client
        st.session_state.pathao_client_secret = new_p_secret
        st.session_state.pathao_email = new_p_email
        st.session_state.pathao_password = new_p_pass
        st.session_state.pathao_store_id = new_p_store
        if new_custom_prod: learn_new_product(new_custom_prod.strip())
        
        sync_profile_to_db()
        st.success("✅ Settings Saved Successfully to Supabase!")
        
    st.markdown("---")
    if st.button("Reset Memory / Clear App Data", type="secondary"):
        st.session_state.all_orders, st.session_state.ignored_messages = [], []
        st.session_state.total_extracted_today = 0
        st.session_state.total_scanned = 0
        st.session_state.chat_history = [{"role": "assistant", "content": "হ্যালো! আমি NWOP এআই। আপনার বিজনেস ডাটা বা অন্য যেকোনো বিষয়ে আমি সাহায্য করতে প্রস্তুত।"}]
        log_task("App memory completely wiped.")
        st.rerun()

with tab_about:
    st.header("ℹ️ About Developer")
    st.markdown("---")
    col_a1, col_a2 = st.columns([1, 3])
    with col_a1:
        img_bytes = get_image_bytes("logo.png")
        if img_bytes: st.image(img_bytes, width=150)
        else: st.markdown("<h1 style='font-size: 80px; margin-top: -20px;'>👨‍💻</h1>", unsafe_allow_html=True)
    with col_a2:
        st.markdown("### **Nazrul's Whatsapp Order Parser (NWOP)**")
        st.write("An enterprise-grade automation tool designed to extract, parse, and manage WhatsApp orders with high accuracy, smart formatting, duplicate detection, and direct Excel compilation.")
    
    st.markdown("#### 👨‍💻 Developer Profile")
    st.markdown("""
    * **Name:** Nazrul Rana
    * **Education:** B.Sc. in Computer Science and Engineering (CSE), Southeast University.
    * **WhatsApp:** 01641434000
    * **Facebook:** [nazrulranaxD.s](https://www.facebook.com/nazrulranaxD.s)
    * **LinkedIn:** [nazrulhuda](https://www.linkedin.com/in/nazrulhuda/)
    * **Version:** 52.1 (The Final UI & Sync Fix)
    """)
    st.info("For any bug reports, feature requests, custom automation tools, or software development inquiries, please feel free to reach out via WhatsApp.")
